from datetime import datetime
from openpyxl import load_workbook
from config import EXCEL_PATH
from database import get_connection


def generar_id_sesion(codigo_equipo):
    ahora = datetime.now()
    return f"SES-{codigo_equipo}-{ahora.strftime('%Y%m%d-%H%M%S')}"


def guardar_sesion_sqlite(sesion, detalles):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute(
            """
            INSERT INTO sesiones_verificacion (
                id_sesion,
                codigo_equipo,
                nombre_equipo,
                laboratorio,
                fecha,
                hora,
                responsable,
                estado,
                total_puntos,
                puntos_cumplen,
                puntos_no_cumplen,
                puntos_no_evaluados
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                sesion["id_sesion"],
                sesion["codigo_equipo"],
                sesion["nombre_equipo"],
                sesion["laboratorio"],
                sesion["fecha"],
                sesion["hora"],
                sesion["responsable"],
                sesion["estado"],
                sesion["total_puntos"],
                sesion["puntos_cumplen"],
                sesion["puntos_no_cumplen"],
                sesion["puntos_no_evaluados"],
            ),
        )

        for detalle in detalles:
            cur.execute(
                """
                INSERT INTO detalle_verificacion (
                    id_sesion,
                    codigo_equipo,
                    punto,
                    nombre_chequeo,
                    valor_nominal,
                    resultado,
                    error,
                    limite_inferior,
                    limite_superior,
                    estado_punto,
                    observacion
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    sesion["id_sesion"],
                    detalle["codigo_equipo"],
                    detalle["punto"],
                    detalle["nombre_chequeo"],
                    detalle["valor_nominal"],
                    detalle["resultado"],
                    detalle["error"],
                    detalle["limite_inferior"],
                    detalle["limite_superior"],
                    detalle["estado_punto"],
                    detalle["observacion"],
                ),
            )

            if detalle["estado_punto"] != "Cumple" or detalle["observacion"] != "Sin novedades":
                cur.execute(
                    """
                    INSERT INTO bitacora (
                        fecha,
                        hora,
                        codigo_equipo,
                        evento,
                        detalle,
                        usuario,
                        origen
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        sesion["fecha"],
                        sesion["hora"],
                        sesion["codigo_equipo"],
                        "Evento en verificación",
                        f"{detalle['punto']} - {detalle['estado_punto']} - {detalle['observacion']}",
                        sesion["responsable"],
                        "Verificación",
                    ),
                )

        conn.commit()
        ok,msg=guardar_verificaciones_excel(sesion,detalles)
        if not ok:
            return False,f"SQLite OK, Excel: {msg}"
        return True,"Sesión guardada correctamente en SQLite y Excel."

    except Exception as e:
        conn.rollback()
        return False, f"Error guardando sesión: {e}"

    finally:
        conn.close()

def guardar_verificaciones_excel(sesion, detalles):
    if not EXCEL_PATH.exists():
        return False, f"No existe {EXCEL_PATH}"
    libro=load_workbook(EXCEL_PATH)
    if "Verificaciones" not in libro.sheetnames:
        return False,"No existe hoja Verificaciones"
    ws=libro["Verificaciones"]
    headers=[c.value for c in ws[1]]
    idx={str(v).strip():i+1 for i,v in enumerate(headers) if v}
    def w(r,n,v):
        if n in idx: ws.cell(row=r,column=idx[n],value=v)
    for d in detalles:
        r=ws.max_row+1
        vals={
        "id_verificacion":sesion["id_sesion"],"id_sesion":sesion["id_sesion"],
        "fecha_verificacion":sesion["fecha"],"codigo_equipo":sesion["codigo_equipo"],
        "nombre_equipo":sesion["nombre_equipo"],"laboratorio":sesion["laboratorio"],
        "nombre_chequeo":d["nombre_chequeo"],"codigo_patron":d.get("codigo_patron",""),
        "fecha_vencimiento_patron":d.get("fecha_vencimiento_patron",""),
        "valor_esperado_g":d["valor_nominal"],"valor_observado_g":d["resultado"],
        "desviacion_g":d["error"],"limite_inferior_g":d["limite_inferior"],
        "limite_superior_g":d["limite_superior"],"cumple":d["estado_punto"],
        "responsable":sesion["responsable"],"observaciones":d["observacion"],
        "estado_sesion":sesion["estado"],"usuario_login":sesion["responsable"],
        "fecha_registro":datetime.now().strftime("%Y-%m-%d %H:%M:%S"),"origen":"PROVICHECK"}
        for k,v in vals.items(): w(r,k,v)
    libro.save(EXCEL_PATH); libro.close()
    return True,"OK"